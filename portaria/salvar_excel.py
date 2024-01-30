from flask import Flask, request
from openpyxl import Workbook, load_workbook

app = Flask(__name__)

@app.route('/salvar_excel.py', methods=['POST'])
def salvar_em_excel():
    dados = request.json

    nome_arquivo = "bancodedados.xlsx"
    
    # Se o arquivo Excel não existir, crie um novo e adicione um cabeçalho
    try:
        workbook = load_workbook(nome_arquivo)
    except FileNotFoundError:
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["Nome", "Cargo", "Centro de Custo", "N° do Crachá", "CPF"])
    
    sheet = workbook.active

    # Adicione os dados ao Excel
    nova_linha = [dados["nome"], dados["cargo"], dados["centroCusto"], dados["cracha"], dados["cpf"]]
    sheet.append(nova_linha)

    # Salve o arquivo Excel
    workbook.save(nome_arquivo)

    return "Dados salvos com sucesso no Excel!"

if __name__ == '__main__':
    app.run()
